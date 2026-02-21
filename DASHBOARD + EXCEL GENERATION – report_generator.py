from openpyxl import load_workbook
from openpyxl.chart import PieChart, BarChart, Reference
import pandas as pd

def generate_report(final_df, output_file="GST_Reconciliation_Report.xlsx"):

    final_df.to_excel(output_file, sheet_name="Reconciliation", index=False)

    wb = load_workbook(output_file)

    # ---------------- DASHBOARD DATA ----------------
    total_2b = final_df["SUPPLIER GSTIN_2B"].notna().sum()
    total_books = final_df["SUPPLIER GSTIN_BOOKS"].notna().sum()

    status_counts = final_df["MATCH STATUS"].value_counts()

    dashboard = wb.create_sheet("Dashboard")

    dashboard["A1"] = "Total Records in 2B"
    dashboard["B1"] = total_2b

    dashboard["A2"] = "Total Records in Books"
    dashboard["B2"] = total_books

    row = 4
    for status, count in status_counts.items():
        dashboard[f"A{row}"] = status
        dashboard[f"B{row}"] = count
        row += 1

    # Pie Chart
    pie = PieChart()
    data = Reference(dashboard, min_col=2, min_row=4, max_row=row-1)
    labels = Reference(dashboard, min_col=1, min_row=4, max_row=row-1)
    pie.add_data(data, titles_from_data=False)
    pie.set_categories(labels)
    pie.title = "Match Status Distribution"

    dashboard.add_chart(pie, "D4")

    wb.save(output_file)
